import os, base64, json, uuid, io, re, hashlib
from typing import Optional, List
from flask import Blueprint, jsonify, request, render_template, send_from_directory, current_app
from sqlalchemy import text
import boto3
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from secrets import token_bytes
from app import db
from routes_auth import jwt_required  # 인증 데코레이터

# 파일 파싱 라이브러리
import pdfplumber
import docx
import openpyxl
import pandas as pd

# ─────────────────────────────────────────────
# 환경설정
# ─────────────────────────────────────────────
REGION = os.getenv("AWS_REGION", "ap-northeast-2")
KMS_KEY_ID = os.getenv("KMS_KEY_ID", "alias/KEK")

crypto_bp = Blueprint("crypto", __name__)
kms = boto3.client("kms", region_name=REGION)

# ─────────────────────────────────────────────
# 유틸 함수
# ─────────────────────────────────────────────
def sha256_hex(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()

def parse_any(filename: str, b: bytes) -> str:
    fn = filename.lower()
    try:
        if fn.endswith(".pdf"):
            with pdfplumber.open(io.BytesIO(b)) as pdf:
                return "\n".join((p.extract_text() or "") for p in pdf.pages)
        if fn.endswith(".docx"):
            d = docx.Document(io.BytesIO(b))
            return "\n".join(p.text for p in d.paragraphs if p.text)
        if fn.endswith(".xlsx"):
            wb = openpyxl.load_workbook(io.BytesIO(b), data_only=True)
            out = []
            for s in wb.sheetnames:
                for row in wb[s].iter_rows(values_only=True):
                    out.append(" ".join(str(c) for c in row if c))
            return "\n".join(out)
        if fn.endswith(".csv"):
            try:
                return pd.read_csv(io.BytesIO(b), encoding="utf-8").to_string(index=False)
            except Exception:
                return pd.read_csv(io.BytesIO(b), encoding="cp949").to_string(index=False)
        if fn.endswith(".json"):
            return json.dumps(json.loads(b.decode("utf-8", "ignore")), ensure_ascii=False, indent=2)
        if fn.endswith(".txt"):
            return b.decode("utf-8", "ignore")
    except Exception:
        return b.decode("utf-8", "ignore")
    return ""

# ─────────────────────────────────────────────
# PII 마스킹 규칙
# ─────────────────────────────────────────────
AUTH_VALUE_SUB_RE = re.compile(
    r'((password|api[_ ]?key|token)\s*[:=]\s*)([\'"]?)[^\'",\s]+([\'"]?)',
    re.IGNORECASE
)

PII_RULES_MAP = {
    # 주민등록번호
    "rrn": (
        re.compile(r"\b\d{6}-\d{7}\b"),
        lambda s: "######-*******"
    ),
    # 연락처/전화/휴대폰 라벨 허용 + 전각 콜론 허용
    "phone": (
        re.compile(r"(?:연락처|전화|휴대폰)?\s*[:：]?\s*01[016789][-\s]?\d{3,4}[-\s]?\d{4}"),
        lambda s: (lambda d: (d[:3] + "-****-" + d[-4:]) if len(d) >= 10 else "010-****-****")(re.sub(r"\D", "", s))
    ),
    # 이메일
    "email": (
        re.compile(r"\b([A-Za-z0-9._%+-])([A-Za-z0-9._%+-]*)(@[A-Za-z0-9.-]+\.[A-Za-z]{2,})\b"),
        lambda s: (s[0] + "***" + s[s.find('@'):]) if "@" in s else s
    ),
    # 10~19자리 연속 숫자(카드/계좌)
    "card_or_acct": (
        re.compile(r"\b\d{10,19}\b"),
        lambda s: (s[:6] + "******" + s[-4:]) if len(s) >= 10 else "******"
    ),
    # 비밀번호/토큰/API Key 등
    "auth": (
        re.compile(r"\b(password|api[_ ]?key|token)\s*[:=]\s*['\"]?[^'\",\s]+['\"]?", re.IGNORECASE),
        lambda s: AUTH_VALUE_SUB_RE.sub(r"\1[SECRET]", s)
    ),
    # 시/도 + 구/군/시까지만 남기고 이후 마스킹
    "address": (
        re.compile(r"[가-힣]+(?:시|도)\s?[가-힣]+(?:구|군|시)[^\n]*"),
        lambda s: re.sub(r"([가-힣]+(?:시|도)\s?[가-힣]+(?:구|군|시)).*", r"\1 ****", s)
    ),
    # 여권번호: 영문 1자리 + 숫자 8자리
    "passport": (
        re.compile(r"\b[A-Z][0-9]{8}\b", re.IGNORECASE),
        lambda s: s[0] + "********"
    ),
    # 운전면허번호(대표 포맷 두 가지)
    "license": (
        re.compile(r"\b\d{2}-\d{2}-\d{6}-\d{2}\b|\b\d{2}-\d{6}-\d{2}-\d{2}\b"),
        lambda s: re.sub(r"\d", "*", s)
    ),
    # 외국인등록번호(뒤 첫 자리 5~8)
    "foreign_id": (
        re.compile(r"\b\d{6}-[5-8]\d{6}\b"),
        lambda s: "######-*******"
    ),
}

# 항상 마스킹(선택과 무관하게 적용)
ALWAYS_MASK = {"rrn", "passport", "license", "foreign_id", "auth"}

# ─────────────────────────────────────────────
# allowed_types 정규화(문자열/리스트/별칭 대응)
# ─────────────────────────────────────────────
VALID_KEYS = set(PII_RULES_MAP.keys())
ALIASES = {
    "tel": "phone", "phone_number": "phone", "연락처": "phone", "휴대폰": "phone", "전화": "phone",
    "rrn": "rrn", "주민번호": "rrn",
    "mail": "email", "이메일": "email",
    "card": "card_or_acct", "account": "card_or_acct", "계좌": "card_or_acct", "카드": "card_or_acct",
    "auth": "auth", "secret": "auth", "token": "auth", "apikey": "auth",
    "addr": "address", "주소": "address",
}

def _normalize_allowed_types(allowed_types):
    # None/빈 값 → 전체
    if not allowed_types:
        return list(VALID_KEYS)

    if isinstance(allowed_types, str):
        parts = [p.strip() for p in allowed_types.split(",") if p.strip()]
    else:
        parts = []
        for x in allowed_types:
            if x is None:
                continue
            s = str(x)
            if "," in s:
                parts.extend([p.strip() for p in s.split(",") if p.strip()])
            else:
                parts.append(s.strip())

    out = []
    for p in parts:
        k = ALIASES.get(p.lower(), p.lower())
        if k in VALID_KEYS:
            out.append(k)
    return out

# ─────────────────────────────────────────────
# PII 처리
# ─────────────────────────────────────────────
def process_pii(text: str, allowed_types: Optional[List[str]] = None):
    """
    allowed_types:
      - None  -> 전체(VALID_KEYS) 적용
      - list  -> 지정된 타입만 적용
    """
    hits, stats, masked = [], {}, text

    types_to_use = _normalize_allowed_types(allowed_types)
    if not types_to_use:
        # 무효 키만 들어온 경우 전체 적용(사일런트 실패 방지)
        types_to_use = list(PII_RULES_MAP.keys())

    for label in types_to_use:
        pattern, mask_fn = PII_RULES_MAP[label]

        def repl(m):
            orig = m.group(0)
            masked_val = mask_fn(orig)
            hits.append({
                "type": label,
                "masked": masked_val,
                "sha256": sha256_hex(orig.encode("utf-8", "ignore")),
            })
            stats[label] = stats.get(label, 0) + 1
            return masked_val

        masked = pattern.sub(repl, masked)

    return masked, hits, stats

# ─────────────────────────────────────────────
# API
# ─────────────────────────────────────────────
@crypto_bp.get("/api/mask-types")
def list_mask_types():
    return jsonify({"ok": True, "types": sorted(PII_RULES_MAP.keys())}), 200

@crypto_bp.post("/api/encrypt")
@jwt_required
def encrypt():
    try:
        if "file" not in request.files:
            return jsonify({"ok": False, "error": "file_required"}), 400

        f = request.files["file"]
        raw = f.read()
        filename = (f.filename or "upload.bin").strip() or "upload.bin"
        name, ext = os.path.splitext(filename)
        token = str(uuid.uuid4())

        # --- 프론트에서 원하는 마스킹 타입 가져오기 ---
        mask_types = None
        if request.form.get("mask_types"):
            try:
                mask_types = json.loads(request.form["mask_types"])
            except Exception:
                mask_types = request.form.getlist("mask_types")
        elif request.is_json:
            body = request.get_json(silent=True) or {}
            mask_types = body.get("mask_types")

        # 🔧 최종 타입 계산
        # - 사용자가 아무것도 안 보냈으면: 전체(VALID_KEYS) 적용 → process_pii(None)
        # - 일부만 보냈으면: ALWAYS_MASK ∪ 요청값
        if not mask_types:
            types_to_use = None  # process_pii(None) → 전체 적용
        else:
            norm = _normalize_allowed_types(mask_types)
            types_to_use = sorted(ALWAYS_MASK | set(norm)) or None

        # 마스킹 처리
        parsed_text = parse_any(filename, raw)
        if parsed_text:
            masked_text, _, stats = process_pii(parsed_text, types_to_use)
            masked_bytes = masked_text.encode("utf-8", "ignore")
        else:
            masked_bytes = b"*" * len(raw)
            stats = {}

        # 마스킹 파일 저장
        masked_dir = os.path.join("static", "masked")
        os.makedirs(masked_dir, exist_ok=True)
        masked_name = f"{(name or 'upload')}.mask{ext or '.bin'}"
        with open(os.path.join(masked_dir, masked_name), "wb") as out:
            out.write(masked_bytes)

        # KMS Data Key 생성
        resp = kms.generate_data_key(KeyId=KMS_KEY_ID, KeySpec="AES_256")
        dek_plain = resp["Plaintext"]
        wdek = resp["CiphertextBlob"]

        # AES-GCM 암호화
        iv = token_bytes(12)
        aesgcm = AESGCM(dek_plain)
        ciphertext = aesgcm.encrypt(iv, raw, None)

        # DB 저장
        meta = {
            "filename": filename,
            "pii_count": sum(stats.values()) if stats else 0,
            "pii_stats": stats,
            "masked_file": f"/upload/masked/{masked_name}",
        }
        sql = text("""
            INSERT INTO crypto_store.crypto_tokens
                (token, wdek, iv, ciphertext, tag, meta)
            VALUES
                (:token, :wdek, :iv, :ciphertext, :tag, CAST(:meta AS JSONB))
        """)
        db.session.execute(sql, {
            "token": token,
            "wdek": base64.b64encode(wdek).decode(),
            "iv": base64.b64encode(iv).decode(),
            "ciphertext": base64.b64encode(ciphertext).decode(),
            "tag": "",  # AESGCM.encrypt 결과에 tag 포함
            "meta": json.dumps(meta, ensure_ascii=False),
        })
        db.session.commit()

        # 실제 적용 규칙을 응답에 포함(프론트 표시용)
        effective_mask_types = sorted(VALID_KEYS) if types_to_use is None else list(types_to_use)

        return jsonify({
            "ok": True,
            "token": token,
            "masked_file_url": f"/upload/masked/{masked_name}",
            "pii_stats": stats,
            "masked_preview": masked_bytes[:500].decode("utf-8", "ignore"),
            "effective_mask_types": effective_mask_types,  # 👈 프론트에서 rulesText 계산 시 사용
        }), 200

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("encrypt_failed")
        return jsonify({"ok": False, "error": "encrypt_failed", "message": str(e)}), 500

@crypto_bp.post("/api/decrypt")
@jwt_required
def decrypt():
    try:
        body = request.get_json(force=True, silent=True) or {}
        token = body.get("token")
        if not token:
            return jsonify({"ok": False, "error": "missing_token"}), 400

        row = db.session.execute(text("""
            SELECT wdek, iv, ciphertext, meta
            FROM crypto_store.crypto_tokens
            WHERE token = :token
        """), {"token": token}).fetchone()
        if not row:
            return jsonify({"ok": False, "error": "token_not_found"}), 404

        wdek_b64, iv_b64, ct_b64, raw_meta = row

        wdek = base64.b64decode(wdek_b64)
        iv = base64.b64decode(iv_b64)
        ciphertext = base64.b64decode(ct_b64)

        # 메타
        meta = {}
        if isinstance(raw_meta, str):
            try:
                meta = json.loads(raw_meta)
            except Exception:
                pass

        # KMS 복호화
        resp = kms.decrypt(CiphertextBlob=wdek)
        dek_plain = resp["Plaintext"]

        # AES-GCM 복호화
        aesgcm = AESGCM(dek_plain)
        plaintext = aesgcm.decrypt(iv, ciphertext, None)

        return jsonify({
            "ok": True,
            "result": base64.b64encode(plaintext).decode(),
            "filename": meta.get("filename", "decrypted_file"),
        }), 200

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("decrypt_failed")
        return jsonify({"ok": False, "error": "decrypt_failed", "message": str(e)}), 500

# 업로드 페이지
@crypto_bp.route("/upload/")
def upload_page():
    return render_template("upload.html")

# 마스킹 파일 다운로드
@crypto_bp.get("/upload/masked/<path:fname>")
def download_masked(fname):
    return send_from_directory("static/masked", fname, as_attachment=True)

